import logging as log
import pandas as pd
from openpyxl import load_workbook
import datetime

class Juego:
    def __init__(self, numeroMaximoIntentos):
        self.numeroMaximoIntentos = numeroMaximoIntentos

    def num(self):
        log.info(self.numeroMaximoIntentos)


class JuegoNumero(Juego):
    def __init__(self, numeroMaximoIntentos, longitudPalabra):
        super().__init__(numeroMaximoIntentos)
        self.numero = longitudPalabra
        self.intentoActual = 0
        self.aciertoLongitud = False
        self.intentosNumeroLetras = []

    def adivinarLongitudPalabra(self):
        correcto = False
        self.intentosNumeroLetras.append(self.intentoActual)
        if self.numero == self.intentoActual:
            self.aciertoLongitud = True
            return self.aciertoLongitud
        else:
            self.numeroMaximoIntentos -= 1
            if self.numeroMaximoIntentos <= 0:
                return self.aciertoLongitud
            else:
                self.intentosMaximosRestantes()
                if self.numero < self.intentoActual:
                    while not correcto:
                        self.intentoActual = input('Introduce un número menor:\n')
                        try:
                            self.intentoActual = int(self.intentoActual)
                            correcto = True
                            return self.aciertoLongitud
                        except:
                            log.warning("Solo se admiten números")

                elif self.numero > self.intentoActual:
                    while not correcto:
                        self.intentoActual = input('Introduce un número mayor:\n')
                        try:
                            self.intentoActual = int(self.intentoActual)
                            correcto = True
                            return self.aciertoLongitud
                        except:
                            log.warning("Solo se admiten números")

    def intentosMaximosRestantes(self):
        log.info("Te quedan solo " + str(self.numeroMaximoIntentos) + " intentos")


class JuegoPalabra(JuegoNumero):
    def __init__(self, numeroMaximoIntentos, palabraAdivinar, longitudPalabra):
        super().__init__(numeroMaximoIntentos, longitudPalabra)
        self.palabraAdivinar = palabraAdivinar
        self.longitudPalabra = longitudPalabra
        self.intentosPalabras = []
        self.intentosLetras = []
        self.aciertos = []
        self.palabraLista = list(self.palabraAdivinar.upper())
        self.victoria = False
        self.final = False

        for x in self.palabraAdivinar:
            self.aciertos.append('#')

    def comprobarAcierto(self, letra):
        letras = []
        for posicion, caracter in enumerate(self.palabraAdivinar.upper()):
            if caracter == letra:
                letras.append(posicion)

        for x in letras:
            self.aciertos[x] = letra

    def dibujarPalabra(self):
        for x in self.aciertos:
            print(x, end='')

    def adivinarLetra(self):
        global letra
        correcta = False
        while not correcta:
            letra = input("\nIntroduzca una letra de la palabra\n")
            if letra.isalpha():
                if len(letra) == 1:
                    letra = str(letra).upper()
                    self.intentosLetras.append(letra)
                    correcta = True
                else:
                    log.warning("\nSolo se admite UNA letra\n")
            else:
                log.warning("\nPor favor introduce una cadena de texto\n")

        if letra in self.palabraLista:
            log.info("\nEnhorabuena acertaste una letra")
            self.comprobarAcierto(letra)
            self.dibujarPalabra()
        else:
            self.numeroMaximoIntentos -= 1
            if self.numeroMaximoIntentos != 0:
                self.intentosMaximosRestantes()
                self.dibujarPalabra()
            else:
                self.final = True

    def adivinarPalabra(self):
        global palabra
        correcta = False
        while not correcta:
            palabra = input("Introduzca la palabra ")
            if palabra.isalpha():
                palabra = str(palabra).upper()
                self.intentosPalabras.append(palabra)
                correcta = True
            else:
                log.warning("\nPor favor introduce una cadena de texto\n")
        if palabra == self.palabraAdivinar.upper():
            self.final = True
            self.victoria = True
        else:
            self.numeroMaximoIntentos -= 1
            if self.numeroMaximoIntentos != 0:
                self.intentosMaximosRestantes()
                self.dibujarPalabra()
            else:
                self.final = True

    def resumen(self):
        log.info('Tus intentos de longitud de palabra fueron ' + str(self.intentosNumeroLetras))
        log.info('Tus intentos de letras fueron ' + str(self.intentosLetras))
        log.info('Tus intentos de palabras fueron ' + str(self.intentosPalabras))

    def resultados(self):
        if self.victoria:
            log.info('Felicidades adivinaste la palabra')
            log.info(self.palabraAdivinar)
            self.resumen()
        else:
            log.info("La palabra a adivinar era " + self.palabraAdivinar)
            self.resumen()


class Jugador:
    def __init__(self, nombre, apellido1, apellido2):
        self.nombre = nombre
        self.apellido1 = apellido1
        self.apellido2 = apellido2


class LogJuego:
    def __init__(self, idPersona, idJuego, fechaJuego, resultado):
        self.idPersona = idPersona
        self.idJuego = idJuego
        self.fechaJuego = fechaJuego
        self.resultado = resultado


if __name__ == '__main__':
    log.basicConfig(level=log.INFO)
    usuarios = pd.read_excel(io="usuarios.xlsx", sheet_name="Hoja1")
    dfUsuarios = pd.DataFrame(usuarios)

    juegos = pd.read_excel(io="juegos.xlsx", sheet_name="Hoja1")
    dfJuegos = pd.DataFrame(juegos)

    seguir = False
    while not seguir:
        dni = input("introduzca su dni para poder jugar, sin guiones y todo junto\n")
        if len(dni) != 9:
            log.warning("El DNI consta de 9 caracteres, 8 NUMERO y 1 LETRA al final")
        else:
            letra = dni[-1]
            if letra.isalpha():
                numero = dni[:8]
                dni = numero + "-" + letra
                seguir = True
            else:
                log.warning("El DNI consta de 9 caracteres, 8 NUMERO y 1 LETRA al final")

    claveJuego = 0
    resultado = 1

    if dfUsuarios.isin([dni]).any().any():
        nombre = dfUsuarios[dfUsuarios['dni'] == dni]['nombre']
        apellido1 = dfUsuarios[dfUsuarios['dni'] == dni]['apellido_1']
        apellido2 = dfUsuarios[dfUsuarios['dni'] == dni]['apellido_2']
        player = Jugador(nombre, apellido1, apellido2)
        log.info("Bienvenido " + player.nombre + " " + player.apellido1 + " " + player.apellido2)

        seguir = False
        while not seguir:
            log.info("\nEscoge el juego al que deseas jugar:\n1.Numero\n2.Palabra\n3.Frase")
            opcion = input("\nIntroduce la opcion:\n")
            try:
                opcion = int(opcion)
                seguir = True
            except:
                log.warning("Solo se admiten números entre las opciones disponibles")

        if opcion == 1:
            nombreJuego = "numero"
            claveJuego = dfJuegos[dfJuegos['juego'] == nombreJuego]['id']
        elif opcion == 2:
            nombreJuego = "palabra"
            claveJuego = dfJuegos[dfJuegos['juego'] == nombreJuego]['id']
        elif opcion == 3:
            nombreJuego = "frase"
            claveJuego = dfJuegos[dfJuegos['juego'] == nombreJuego]['id']
        else:
            log.info("Introduce una opcion valida")

        seguir = False
        while not seguir:
            numeroMaximoIntentos = input("Introduce la cantidad de intentos que tendrá el jugador\n")
            try:
                if int(numeroMaximoIntentos) > 0:
                    numeroMaximoIntentos = int(numeroMaximoIntentos)
                    seguir = True
                else:
                    log.warning("\nIntroduce un número mayor que 0\n")
            except:
                log.warning("Solo se admiten números")

        cadenaCorrecta = False
        while not cadenaCorrecta:
            palabraAdivinar = input('Introduce la palabra que tendrá que adivinarse:\n')
            if palabraAdivinar.isalpha():
                longitudPalabraAdivinar = len(palabraAdivinar)
                game = JuegoPalabra(numeroMaximoIntentos, palabraAdivinar, longitudPalabraAdivinar)
                cadenaCorrecta = True
            else:
                log.warning("Por favor introduce una cadena de texto")

        seguir = False
        while not seguir:
            longitudPalabra = input('Introduce el número de letras que crees que tiene la palabra\n')
            try:
                longitudPalabra = int(longitudPalabra)
                game.intentoActual = longitudPalabra
                seguir = True
            except:
                log.warning("Solo se admiten números")

        while not game.victoria:
            if game.adivinarLongitudPalabra():
                while not game.final:
                    seguir = False
                    while not seguir:
                        log.info("\n1. Adivinar una letra\n2. Adivinar la palabra completa")
                        opcion = input("\nIntroduce la opcion:\n")
                        try:
                            opcion = int(opcion)
                            seguir = True
                        except:
                            log.warning("Solo se admiten números entre las opciones disponibles")

                    if opcion == 1:
                        game.adivinarLetra()
                    elif opcion == 2:
                        game.adivinarPalabra()
                    else:
                        log.info("Introduce una opcion valida")

            if game.numeroMaximoIntentos <= 0:
                resultado = 0
                log.info("Te quedaste sin intentos, la longitud era " + str(game.longitudPalabra))
                break

        game.resultados()

        idJugador = int(dfUsuarios[dfUsuarios['dni'] == dni]['id'])
        claveJuego = int(claveJuego)

        book = load_workbook('logs_juego.xlsx')
        writer = pd.ExcelWriter('logs_juego.xlsx', engine='openpyxl')
        writer.book = book

        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        idLog = int(writer.sheets['Hoja1'].max_row)
        log = {
            'id': [idLog],
            'id_persona': [idJugador],
            'id_juego': [claveJuego],
            'fecha_juego': [datetime.date.today()],
            'resultado': [resultado]
        }
        df = pd.DataFrame(log)

        df.to_excel(writer, "Hoja1", startrow=writer.sheets['Hoja1'].max_row, header=False, index=False)

        writer.save()

    else:
        log.info("No figuras en la base de datos, lo siento no puedes jugar")
